from flask import Flask, render_template, request, send_from_directory
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import difflib
import logging
import re
import os

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Logging configuration: detailed logs for every action in this app.
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
)
logger = logging.getLogger("reservoir")

OUTPUT_FOLDER = "output"
TEMPLATE_FILE = "Master Templete.xlsx"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)
logger.info("App starting. OUTPUT_FOLDER='%s', TEMPLATE_FILE='%s'", OUTPUT_FOLDER, TEMPLATE_FILE)


def normalize_name(name):
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def load_template_names():
    logger.debug("Loading template names from '%s'", TEMPLATE_FILE)
    wb = load_workbook(TEMPLATE_FILE)
    sheet = wb.active
    template_names = []

    for row in range(4, sheet.max_row + 1):
        value = sheet.cell(row=row, column=2).value
        if value is not None:
            template_names.append(str(value).strip())

    logger.debug("Loaded %d template reservoir names", len(template_names))
    return template_names


def closest_template_name(input_name, template_names, min_ratio=0.55):
    input_norm = normalize_name(input_name)
    if not input_norm:
        return None

    exact_match = next(
        (name for name in template_names if normalize_name(name) == input_norm),
        None
    )
    if exact_match:
        return exact_match

    best_match = None
    best_score = 0.0

    for template_name in template_names:
        template_norm = normalize_name(template_name)

        if input_norm == template_norm:
            return template_name

        if input_norm in template_norm or template_norm in input_norm:
            return template_name

        score = difflib.SequenceMatcher(
            None,
            input_norm,
            template_norm
        ).ratio()
        if score > best_score:
            best_score = score
            best_match = template_name

    if best_score >= min_ratio:
        return best_match

    if len(input_norm) <= 8 and best_score >= 0.45:
        return best_match

    return None


def find_reservoir_blocks(text, template_names):
    lines = [line.strip() for line in re.split(r'\r?\n', text) if line.strip()]
    blocks = []
    current = None
    reservoir_marker = re.compile(r'\b(?:reservoir|tank)\b', re.I)

    for line in lines:
        raw_line = re.sub(r'^\[.*?\]\s*', '', line).strip()
        if reservoir_marker.search(raw_line):
            candidate = re.sub(r'^\+?[\d\s\-().]*:?', '', raw_line).strip()
            candidate = re.sub(r'^(?:Sir[\s,:-]*)?', '', candidate, flags=re.I).strip()
            template_name = closest_template_name(candidate, template_names)
            if template_name:
                current = {'name': template_name, 'lines': [line]}
                blocks.append(current)
                continue
        if current:
            # stop block when a new timestamped message begins without a reservoir line
            if re.match(r'^\[\d{1,2}/\d{1,2},', line):
                current = None
                continue
            current['lines'].append(line)

    return blocks


def extract_values_from_block(block_text, template_name):
    block_text = block_text.replace('�', '')

    if template_name == 'Mylaram Balancing Reservoir':
        m = re.search(
            r'present\s*capacity\s*[:\-]?\s*([\d.]+)\s*TMC.*?\+?([\d.]+)',
            block_text,
            re.I | re.S
        )
        if m:
            level = float(m.group(2))
            storage = float(m.group(1)) * 1000
            return level, storage

    level_match = re.search(r'present\s*level\s*[:\-]?\s*\+?([\d.]+)', block_text, re.I)
    storage_match = re.search(r'present\s*storage\s*[:\-]?\s*\+?([\d.]+)', block_text, re.I)

    if not storage_match:
        storage_match = re.search(r'present\s*level.*?\((\d+[\d.]*?)\s*mcft\)', block_text, re.I | re.S)

    if level_match and storage_match:
        return float(level_match.group(1)), float(storage_match.group(1))

    # fallback: parse present capacity as storage if no explicit present storage line exists
    fallback = re.search(
        r'present\s*capacity\s*[:\-]?\s*([\d.]+)\s*TMC.*?\+?([\d.]+)',
        block_text,
        re.I | re.S
    )
    if fallback:
        return float(fallback.group(2)), float(fallback.group(1)) * 1000

    return None, None


def extract_reservoir_entries(text, template_names):
    entries = {}

    blocks = find_reservoir_blocks(text, template_names)
    logger.debug("Found %d reservoir block(s) in message", len(blocks))
    for block in blocks:
        block_text = '\n'.join(block['lines'])
        level, storage = extract_values_from_block(block_text, block['name'])
        if level is not None and storage is not None:
            logger.info("Parsed '%s': level=%s, storage=%s", block['name'], level, storage)
            entries[block['name']] = [level, storage]
        else:
            logger.warning("Could not extract values for '%s'", block['name'])

    return entries


def parse_messages(text):
    data = {}
    date = ""

    logger.info("Parsing message (%d chars)", len(text or ""))
    date_match = re.search(r'(\d{2}[.-]\d{2}[.-]\d{4})', text)
    if date_match:
        date = date_match.group(1).replace("-", ".")
    logger.info("Date detected: '%s'", date)

    template_names = load_template_names()
    data.update(extract_reservoir_entries(text, template_names))

    logger.info("Parsing complete: %d reservoir(s) with data", len(data))
    return date, data


@app.route("/", methods=["GET", "POST"])
def home():

    table_html = ""
    missing_html = ""
    date = ""
    message = ""
    preview_status = ""
    download_file = ""
    data_count = 0
    missing_count = 0

    logger.info("%s / request received", request.method)

    if request.method == "POST":

        action = request.form.get("action")

        message = request.form.get("message", "")
        logger.info("Action='%s' submitted; message length=%d", action, len(message))

        date, data = parse_messages(message)

        # Load template names for comparison
        template_names = load_template_names()

        for reservoir, values in data.items():

            table_html += f"""
            <tr>
                <td>{reservoir}</td>
                <td>{values[0]}</td>
                <td>{values[1]}</td>
            </tr>
            """

        data_count = len(data)

        # Find missing reservoirs
        missing_reservoirs = [r for r in template_names if r not in data]
        missing_count = len(missing_reservoirs)
        logger.info("Preview built: %d with data, %d missing (no data)", data_count, missing_count)
        if missing_reservoirs:
            logger.debug("Missing reservoirs: %s", ", ".join(missing_reservoirs))

        for reservoir in missing_reservoirs:
            missing_html += f"""
            <tr>
                <td>{reservoir}</td>
                <td colspan="2" style="text-align: center; color: #9b1c1c; font-weight: 600;">No data</td>
            </tr>
            """

        if not data:
            logger.warning("No reservoir data found in the input message")
            preview_status = (
                "No reservoir data found in the input. "
                "Please check the message content and reservoir names."
            )

        if action == "generate":
            logger.info("Generate Excel started for date='%s'", date)

            wb = load_workbook("Master Templete.xlsx")

            sheet = wb.active

            sheet["A1"] = (
                f"DAILY WATER LEVELS IN RESERVOIRS "
                f"UNDER IRRIGATION CIRCLE, JANGAON\nDATED: {date}"
            )

            # Define fill colors for remarks based on percentage
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

            for row in range(4, sheet.max_row + 1):

                reservoir = sheet[f'B{row}'].value

                if reservoir is None:
                    continue

                excel_name = str(reservoir).strip()

                if excel_name in data:
                    level = data[excel_name][0]
                    storage = data[excel_name][1]
                    sheet[f'G{row}'] = level
                    sheet[f'H{row}'] = storage
                    logger.info("Row %d '%s': G(level)=%s, H(storage MCFT)=%s", row, excel_name, level, storage)
                else:
                    sheet[f'G{row}'] = "NA"
                    sheet[f'H{row}'] = "NA"
                    logger.debug("Row %d '%s': no data -> NA", row, excel_name)

                # Calculate Gross capacity in TMC from storage in H
                h_value = sheet[f'H{row}'].value
                if h_value is not None and str(h_value) != "NA":
                    try:
                        h_value_float = float(h_value)
                        sheet[f'I{row}'] = round(h_value_float / 1000, 3)
                    except (ValueError, TypeError):
                        sheet[f'I{row}'] = "NA"
                else:
                    sheet[f'I{row}'] = "NA"

                # Calculate Percentage of Filling from H / F
                total_capacity = sheet[f'F{row}'].value
                if h_value is not None and str(h_value) != "NA" and total_capacity is not None and str(total_capacity) != "NA":
                    try:
                        percentage = (float(h_value) / float(total_capacity)) * 100
                        sheet[f'J{row}'] = round(percentage, 2)
                        remarks_cell = sheet[f'K{row}']
                        remarks_cell.value = ""
                        if percentage > 85:
                            remarks_cell.fill = red_fill
                            colour = "red"
                        elif 50 <= percentage <= 85:
                            remarks_cell.fill = yellow_fill
                            colour = "yellow"
                        else:
                            remarks_cell.fill = green_fill
                            colour = "green"
                        logger.info("Row %d '%s': TMC=%s, %%filling=%.2f -> remarks=%s",
                                    row, excel_name, sheet[f'I{row}'].value, percentage, colour)
                    except (ValueError, TypeError, ZeroDivisionError):
                        logger.warning("Row %d '%s': could not compute %% filling", row, excel_name)
                        sheet[f'J{row}'] = "NA"
                else:
                    sheet[f'J{row}'] = "NA"
                    sheet[f'K{row}'] = ""

            filename = f"SE_IC_JGN_MAJOR_WATER_LEVELS_{date}.xlsx"

            filepath = os.path.join(
                OUTPUT_FOLDER,
                filename
            )

            wb.save(filepath)
            logger.info("Excel generated and saved: '%s'", filepath)

            download_file = filename

    return render_template(
        "index.html",
        table=table_html,
        missing_table=missing_html,
        data_count=data_count,
        missing_count=missing_count,
        preview_status=preview_status,
        date=date,
        message=message,
        download_file=download_file
    )


@app.route("/download/<filename>")
def download(filename):
    logger.info("Download requested: '%s'", filename)
    return send_from_directory(
        OUTPUT_FOLDER,
        filename,
        as_attachment=True
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)